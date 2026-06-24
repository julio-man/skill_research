from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook

from skill_research.core.types import SkillRef, Task
from skill_research.executors.spreadsheet_python import SpreadsheetPythonExecutor


class CopyBackend:
    def complete(self, request):
        return "from openpyxl import load_workbook\nwb = load_workbook(INPUT_WORKBOOK)\nwb.save(OUTPUT_WORKBOOK)\n"


def test_spreadsheet_executor_records_relative_paths_and_uses_relative_prelude(tmp_path: Path, monkeypatch) -> None:
    workbook = tmp_path / "input.xlsx"
    Workbook().save(workbook)
    skill = tmp_path / "skill"
    skill.mkdir()
    (skill / "SKILL.md").write_text("# Skill\n", encoding="utf-8")
    monkeypatch.chdir(tmp_path)

    result = SpreadsheetPythonExecutor(CopyBackend()).run(Task("t1", "copy", workbook), SkillRef(skill), tmp_path / "out", {})

    assert not result.artifact_path.startswith("/")
    assert not (result.code_path or "").startswith("/")
    code = Path(result.code_path).read_text(encoding="utf-8")
    assert str(tmp_path) not in code
    assert "INPUT_WORKBOOK = '../input.xlsx'" in code
    assert "OUTPUT_WORKBOOK = 't1_candidate.xlsx'" in code
