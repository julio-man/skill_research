from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook

from skill_research.core.types import SkillRef, Task
from skill_research.executors.spreadsheet_python import SpreadsheetPythonExecutor
from skill_research.llms.replay import ReplayLLMBackend
from skill_research.patches.proposers.openai_trace import OpenAITracePatchProposer
from skill_research.traces.types import TraceRecord


def _workbook(path: Path, value: int) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = value
    wb.save(path)


def test_spreadsheet_executor_accepts_standard_llm_backend(tmp_path: Path) -> None:
    initial = tmp_path / "initial.xlsx"
    _workbook(initial, 1)
    skill_dir = tmp_path / "skill"
    skill_dir.mkdir()
    (skill_dir / "SKILL.md").write_text("# Skill\n", encoding="utf-8")
    backend = ReplayLLMBackend(["from openpyxl import load_workbook\nwb = load_workbook(INPUT_WORKBOOK)\nwb.save(OUTPUT_WORKBOOK)\n"])

    result = SpreadsheetPythonExecutor(backend).run(Task("t1", "copy workbook", initial), SkillRef(skill_dir), tmp_path / "run", {"temperature": 0.0, "max_tokens": 100, "seed": 3})

    assert result.returncode == 0
    assert Path(result.artifact_path).exists()


def test_openai_trace_proposer_accepts_standard_llm_backend(tmp_path: Path) -> None:
    skill_dir = tmp_path / "skill"
    skill_dir.mkdir()
    (skill_dir / "SKILL.md").write_text("# Skill\n", encoding="utf-8")
    backend = ReplayLLMBackend(['{"patches": [{"patch_id": "p1", "patch_type": "add_rule", "target_file": "SKILL.md", "target_section": null, "operation": "append_document", "content": "Rule", "supported_trace_ids": ["t1"]}]}'])

    pool = OpenAITracePatchProposer(backend).propose(SkillRef(skill_dir), [TraceRecord("t1", False, "wrong_answer")], {"k": 1})

    assert pool.patches[0].patch_id == "p1"
