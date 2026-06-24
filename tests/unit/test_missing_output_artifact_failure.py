from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook

from skill_research.core.types import Task
from skill_research.evaluators.spreadsheet import SpreadsheetEvaluator


def test_spreadsheet_evaluator_distinguishes_generated_code_that_did_not_write_artifact(tmp_path: Path) -> None:
    golden = tmp_path / "golden.xlsx"
    Workbook().save(golden)
    task = Task("t1", "copy", metadata={"golden_workbook_path": str(golden), "answer_cells": ["Sheet1!A1"]})
    execution = {
        "artifact_path": str(tmp_path / "missing.xlsx"),
        "returncode": 0,
        "raw_output": "from openpyxl import load_workbook\n# forgot save\n",
        "stderr": "",
    }

    result = SpreadsheetEvaluator().evaluate(task, execution)

    assert result.failure_type == "missing_output_artifact"


def test_spreadsheet_evaluator_distinguishes_runtime_execution_error_without_artifact(tmp_path: Path) -> None:
    golden = tmp_path / "golden.xlsx"
    Workbook().save(golden)
    task = Task("t1", "copy", metadata={"golden_workbook_path": str(golden), "answer_cells": ["Sheet1!A1"]})
    execution = {
        "artifact_path": str(tmp_path / "missing.xlsx"),
        "returncode": 1,
        "raw_output": "from openpyxl import load_workbook\nraise RuntimeError('x')\n",
        "stderr": "RuntimeError: x",
    }

    result = SpreadsheetEvaluator().evaluate(task, execution)

    assert result.failure_type == "execution_error"
