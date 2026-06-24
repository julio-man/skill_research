from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook

from skill_research.core.types import SkillRef, Task
from skill_research.executors.spreadsheet_python import SpreadsheetPythonExecutor


class CopyBackend:
    def complete(self, request):
        return "from openpyxl import load_workbook\nwb = load_workbook(INPUT_WORKBOOK)\nwb.save(OUTPUT_WORKBOOK)\n"


def test_spreadsheet_executor_uses_absolute_input_path_when_running_from_output_dir(tmp_path: Path, monkeypatch) -> None:
    workbook_dir = tmp_path / "relative_data"
    workbook_dir.mkdir()
    workbook_path = workbook_dir / "input.xlsx"
    wb = Workbook()
    wb.save(workbook_path)
    skill_dir = tmp_path / "skill"
    skill_dir.mkdir()
    (skill_dir / "SKILL.md").write_text("# Skill\n", encoding="utf-8")
    monkeypatch.chdir(tmp_path)

    result = SpreadsheetPythonExecutor(CopyBackend()).run(
        Task("t1", "copy", input_path=Path("relative_data/input.xlsx")),
        SkillRef(skill_dir),
        tmp_path / "out",
        {},
    )

    assert result.returncode == 0
    assert Path(result.artifact_path).exists()
