from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook

from skill_research.core.types import SkillRef, Task
from skill_research.datasets.base import DatasetInfo, DatasetSplit
from skill_research.evaluators.base import EvaluationResult
from skill_research.executors.base import ExecutionResult
from skill_research.executors.spreadsheet_python import SpreadsheetPythonExecutor
from skill_research.experiments.benchmark import ComponentBenchmarkRunner
from skill_research.llms.base import CompletionResponse, LLMBackendInfo
from skill_research.llms.openai_backend import OpenAIBackendConfig, OpenAIChatBackend


class FakeUsage:
    def model_dump(self):
        return {"prompt_tokens": 3, "completion_tokens": 4, "total_tokens": 7}


class FakeChoiceMessage:
    content = "ok"


class FakeChoice:
    message = FakeChoiceMessage()


class FakeCompletions:
    def create(self, **kwargs):
        return type("Response", (), {"choices": [FakeChoice()], "model": kwargs["model"], "usage": FakeUsage()})()


class FakeChat:
    def __init__(self):
        self.completions = FakeCompletions()


class FakeClient:
    def __init__(self):
        self.chat = FakeChat()


def test_openai_backend_records_usage_and_elapsed_time() -> None:
    backend = OpenAIChatBackend(OpenAIBackendConfig(model="gpt", api_key="x"), client=FakeClient())

    response = backend.complete(type("Request", (), {"messages": [], "temperature": 0.0, "max_tokens": 10, "seed": None})())

    assert response.metadata["usage"]["total_tokens"] == 7
    assert response.metadata["elapsed_seconds"] >= 0.0


class MetadataBackend:
    def complete(self, request):
        return CompletionResponse("from openpyxl import load_workbook\nwb = load_workbook(INPUT_WORKBOOK)\nwb.save(OUTPUT_WORKBOOK)\n", LLMBackendInfo("fake", "test", "fake"), metadata={"usage": {"total_tokens": 11}, "elapsed_seconds": 0.5})


def test_spreadsheet_executor_records_llm_usage_and_elapsed_time(tmp_path: Path) -> None:
    workbook = tmp_path / "input.xlsx"
    Workbook().save(workbook)
    skill = tmp_path / "skill"
    skill.mkdir()
    (skill / "SKILL.md").write_text("# Skill\n", encoding="utf-8")

    result = SpreadsheetPythonExecutor(MetadataBackend()).run(Task("t1", "copy", workbook), SkillRef(skill), tmp_path / "out", {})

    assert result.metadata["llm_usage"]["total_tokens"] == 11
    assert result.metadata["llm_elapsed_seconds"] == 0.5
    assert result.metadata["executor_elapsed_seconds"] >= 0.0


class FakeExecutor:
    name = "fake"

    def run(self, task, skill, output_dir, config):
        return ExecutionResult("artifact.xlsx", None, "", "", "", 0, metadata={"executor_elapsed_seconds": 0.1})


class FakeEvaluator:
    name = "fake_eval"

    def evaluate(self, task, execution):
        return EvaluationResult(False, 0.0, "wrong_answer")


def test_benchmark_trace_records_evaluation_elapsed_time(tmp_path: Path) -> None:
    split = DatasetSplit("x", [Task("t1", "x")], DatasetInfo("toy", "toy"))
    result = ComponentBenchmarkRunner(split, FakeExecutor(), FakeEvaluator()).run(SkillRef(tmp_path / "skill"), tmp_path / "bench")

    payload = result.traces[0].payload
    assert payload["timing"]["evaluation_elapsed_seconds"] >= 0.0
    assert payload["timing"]["task_total_elapsed_seconds"] >= 0.0
