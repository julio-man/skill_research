from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook

from skill_research.core.types import SkillRef, Task
from skill_research.evaluators.spreadsheet import SpreadsheetEvaluator
from skill_research.executors.spreadsheet_python import SpreadsheetPythonExecutor


def _workbook(path: Path, value: int) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = value
    wb.save(path)


class FakeBackend:
    def complete(self, request):
        return "from openpyxl import load_workbook\nwb = load_workbook(INPUT_WORKBOOK)\nwb['Sheet1']['A1'] = 2\nwb.save(OUTPUT_WORKBOOK)\n"


def test_spreadsheet_python_executor_creates_candidate_workbook(tmp_path: Path) -> None:
    initial = tmp_path / "initial.xlsx"
    _workbook(initial, 1)
    skill_dir = tmp_path / "skill"
    skill_dir.mkdir()
    (skill_dir / "SKILL.md").write_text("# Skill\n", encoding="utf-8")
    task = Task("t1", "Set A1 to 2", input_path=initial, metadata={"golden_workbook_path": str(tmp_path / "golden.xlsx")})

    result = SpreadsheetPythonExecutor(FakeBackend()).run(task, SkillRef(skill_dir), tmp_path / "run", {"temperature": 0.0, "max_tokens": 200})

    assert result.returncode == 0
    assert Path(result.artifact_path).exists()
    assert Path(result.code_path).exists()


def test_spreadsheet_evaluator_compares_answer_cells(tmp_path: Path) -> None:
    candidate = tmp_path / "candidate.xlsx"
    golden = tmp_path / "golden.xlsx"
    _workbook(candidate, 2)
    _workbook(golden, 2)
    task = Task("t1", "Set A1 to 2", metadata={"golden_workbook_path": str(golden), "answer_cells": ["Sheet1!A1"]})

    result = SpreadsheetEvaluator().evaluate(task, {"artifact_path": str(candidate)})

    assert result.passed is True
    assert result.score == 1.0
    assert result.failure_type == "none"
