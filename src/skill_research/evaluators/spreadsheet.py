from __future__ import annotations

from pathlib import Path
import re

from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries

from skill_research.core.types import Task
from skill_research.evaluators.base import CheckResult, EvaluationResult


def _split_regions(value: str) -> list[str]:
    regions = []
    current = []
    in_quote = False
    for char in value:
        if char == "'":
            in_quote = not in_quote
        if char == "," and not in_quote:
            region = "".join(current).strip()
            if region:
                regions.append(region)
            current = []
        else:
            current.append(char)
    region = "".join(current).strip()
    if region:
        regions.append(region)
    return regions


def _parse_region(region: str, default_sheet: str | None) -> tuple[str, str]:
    region = region.strip().strip(",")
    if "!" in region:
        sheet, cells = region.rsplit("!", 1)
        return sheet.strip().strip("'"), cells.strip().strip("'")
    if default_sheet is None:
        raise ValueError(f"No sheet specified for region '{region}'")
    return default_sheet.strip().strip("'"), region.strip().strip("'")


def _expand_cells(sheet: str, cells: str) -> list[str]:
    cells = cells.replace("$", "")
    if ":" not in cells:
        return [f"{sheet}!{cells.upper()}"]
    min_col, min_row, max_col, max_row = range_boundaries(cells)
    refs = []
    for row in range(min_row, max_row + 1):
        for col in range(min_col, max_col + 1):
            from openpyxl.utils.cell import get_column_letter

            refs.append(f"{sheet}!{get_column_letter(col)}{row}")
    return refs


def answer_cell_refs(task: Task, default_sheet: str | None = None) -> list[str]:
    explicit = task.metadata.get("answer_cells")
    if explicit:
        return list(explicit)
    answer_position = task.metadata.get("answer_position")
    if not answer_position:
        return []
    answer_sheet = task.metadata.get("answer_sheet")
    sheets = _split_regions(str(answer_sheet)) if answer_sheet else []
    regions = _split_regions(str(answer_position))
    refs = []
    for index, region in enumerate(regions):
        default_sheet = sheets[index] if index < len(sheets) else (sheets[0] if sheets else default_sheet)
        sheet, cells = _parse_region(region, default_sheet)
        refs.extend(_expand_cells(sheet, cells))
    return refs


class SpreadsheetEvaluator:
    name = "spreadsheet"

    def evaluate(self, task: Task, execution) -> EvaluationResult:
        artifact_path = Path(execution["artifact_path"] if isinstance(execution, dict) else execution.artifact_path)
        golden_path = Path(task.metadata["golden_workbook_path"])
        if not artifact_path.exists():
            return EvaluationResult(False, 0.0, "artifact_missing", [CheckResult("artifact_exists", False, "Artifact missing")])
        candidate = load_workbook(artifact_path, data_only=True)
        golden = load_workbook(golden_path, data_only=True)
        checks: list[CheckResult] = []
        default_sheet = golden.active.title
        for cell_ref in answer_cell_refs(task, default_sheet=default_sheet):
            sheet, cell = cell_ref.split("!", 1)
            expected = golden[sheet][cell].value
            actual = candidate[sheet][cell].value
            passed = expected == actual
            checks.append(CheckResult("cell_value", passed, "passed" if passed else f"{cell_ref}: expected {expected}, got {actual}", {"cell": cell_ref, "expected": expected, "actual": actual}))
        passed = all(check.passed for check in checks)
        score = sum(1 for check in checks if check.passed) / len(checks) if checks else 0.0
        return EvaluationResult(passed, score, "none" if passed else "wrong_answer", checks)
