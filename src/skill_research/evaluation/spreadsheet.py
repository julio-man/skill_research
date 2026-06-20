from __future__ import annotations

import datetime
from pathlib import Path

from openpyxl import load_workbook

from skill_research.data.types import RegionSpec, SpreadsheetTask
from skill_research.evaluation.base import CheckResult, EvaluationResult


class SpreadsheetTaskEvaluator:
    def evaluate(self, task: SpreadsheetTask, candidate_workbook_path: Path) -> EvaluationResult:
        metadata = {
            "task_id": task.task_id,
            "answer_region_raw": task.answer_spec.raw_text,
            "answer_sheet": task.answer_sheet,
        }

        if not candidate_workbook_path.exists():
            return EvaluationResult(
                passed=False,
                score=0.0,
                failure_type="artifact_missing",
                checks=[
                    CheckResult(
                        kind="artifact_exists",
                        passed=False,
                        message="Candidate workbook not found.",
                        details={"candidate_workbook_path": str(candidate_workbook_path)},
                    )
                ],
                metadata=metadata,
            )

        try:
            wb_gold = load_workbook(task.golden_workbook_path, data_only=True)
            wb_candidate = load_workbook(candidate_workbook_path, data_only=True)
        except Exception as exc:
            return EvaluationResult(
                passed=False,
                score=0.0,
                failure_type="artifact_load_error",
                checks=[
                    CheckResult(
                        kind="artifact_load",
                        passed=False,
                        message="Failed to load workbook.",
                        details={"error": str(exc)},
                    )
                ],
                metadata=metadata,
            )

        checks: list[CheckResult] = []
        for region in task.answer_spec.regions:
            result = self._compare_region(wb_gold, wb_candidate, region)
            checks.append(result)
            if not result.passed:
                return EvaluationResult(
                    passed=False,
                    score=0.0,
                    failure_type="wrong_answer",
                    checks=checks,
                    metadata=metadata,
                )

        return EvaluationResult(
            passed=True,
            score=1.0,
            failure_type="none",
            checks=checks,
            metadata=metadata,
        )

    def _compare_region(self, wb_gold, wb_candidate, region: RegionSpec) -> CheckResult:
        sheet_name = region.sheet_name or wb_gold.sheetnames[0]
        if sheet_name not in wb_candidate.sheetnames:
            return CheckResult(
                kind="sheet_exists",
                passed=False,
                message=f"Worksheet '{sheet_name}' not found in candidate workbook.",
                details={"sheet_name": sheet_name},
            )

        ws_gold = wb_gold[sheet_name]
        ws_candidate = wb_candidate[sheet_name]
        cell_names = self._generate_cell_names(region, ws_gold.max_row, ws_gold.max_column)

        for cell_name in cell_names:
            gold_value = ws_gold[cell_name].value
            candidate_value = ws_candidate[cell_name].value
            if not self._compare_cell_value(gold_value, candidate_value):
                return CheckResult(
                    kind="cell_value",
                    passed=False,
                    message=f"Mismatch at {sheet_name}!{cell_name}",
                    details={
                        "sheet_name": sheet_name,
                        "cell": cell_name,
                        "expected": gold_value,
                        "actual": candidate_value,
                    },
                )

        return CheckResult(
            kind="cell_value",
            passed=True,
            message=f"Region matched for {sheet_name}!{region.raw_text}",
            details={"sheet_name": sheet_name, "region": region.raw_text},
        )

    def _generate_cell_names(self, region: RegionSpec, max_row: int, max_column: int) -> list[str]:
        start = region.start_cell
        end = region.end_cell
        if end is None:
            return [start]

        start_col, start_row = self._split_cell(start)
        end_col, end_row = self._split_cell(end)

        if start_row is None and end_row is None:
            start_row = 1
            end_row = max_row
        elif start_row is None or end_row is None:
            raise ValueError("Mixed rowless/ranged cells are not supported")

        start_col_num = self._col_name_to_num(start_col)
        end_col_num = self._col_name_to_num(end_col)
        columns = [self._col_num_to_name(i) for i in range(start_col_num, end_col_num + 1)]
        return [f"{col}{row}" for col in columns for row in range(start_row, end_row + 1)]

    def _split_cell(self, cell: str) -> tuple[str, int | None]:
        col = ""
        row = ""
        for char in cell:
            if char.isdigit():
                row += char
            else:
                col += char
        return col.upper(), int(row) if row else None

    def _col_name_to_num(self, name: str) -> int:
        num = 0
        for char in name:
            num = num * 26 + (ord(char.upper()) - ord("A") + 1)
        return num

    def _col_num_to_name(self, num: int) -> str:
        chars: list[str] = []
        while num > 0:
            num, remainder = divmod(num - 1, 26)
            chars.append(chr(65 + remainder))
        return "".join(reversed(chars))

    def _compare_cell_value(self, left, right) -> bool:
        left = self._normalize_value(left)
        right = self._normalize_value(right)
        if (left == "" and right is None) or (left is None and right == ""):
            return True
        if (left == "" and right == "") or (left is None and right is None):
            return True
        if type(left) is not type(right):
            return False
        return left == right

    def _normalize_value(self, value):
        if isinstance(value, (int, float)):
            return round(float(value), 2)
        if isinstance(value, datetime.time):
            return str(value)[:-3]
        if isinstance(value, datetime.datetime):
            excel_start_date = datetime.datetime(1899, 12, 30)
            delta = value - excel_start_date
            return round(delta.days + delta.seconds / 86400.0, 0)
        if isinstance(value, str):
            try:
                return round(float(value), 2)
            except ValueError:
                return value
        return value
